from openpyxl import Workbook

#new wokbook
wb=Workbook()

#adding multiple sheets
ws1=wb.active
ws1.title="sheet1" #rename the worksheets
ws1['A1']="data for sheet1"

ws2=wb.create_sheet(title='sheet2')
ws2['A1']="data for sheet2"

ws3=wb.create_sheet(title='sheet3')
ws3['A1']="data for sheet3"

wb.save("multiplesheet.xlsx")
